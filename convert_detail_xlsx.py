#!/usr/bin/env python3
"""
轉換客戶訂單(明細分頁) → v8.3 列印格式 + 棧板資訊(含入庫單號)。

新格式(2026-06-26 起)：
  - 讀「明細」分頁，每一列(品號)= 一張貼紙
  - Plant = 倉別(入庫單用)
  - 移倉 = 車趟(如 6/26-1、6/25空調-1)
  - 入庫單號(orderId)：lotte + YYYYMMDD + 車趟代碼 + 流水號(每車趟3碼)
      常溫 6/26-1   → lotte20260626-1001
      恆溫 6/25空調-1 → lotte20260625aircon-1001
  - 排序：車次 > 品號小到大 > 數量小到大
  - 條碼 ean：贈品→單品條碼；其他→外箱條碼
"""

import sys, os, csv, re
try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl")
    import openpyxl

DEFAULT_CUST = "/root/.claude/uploads/cc3263d7-cb46-55b2-8ea0-151cb63cd9f8/cbab5e01-0626__.xlsx"
DEFAULT_MASTER = "/root/.claude/uploads/cc3263d7-cb46-55b2-8ea0-151cb63cd9f8/a0d64486-jtu0612____________20260605__.xlsm"
DEFAULT_OUT = "/home/user/ui-ux-pro-max-skill/ALLY_LOTTE_WH/DATA/0626_棧板資訊"

CUST = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_CUST
MASTER = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_MASTER
OUT_BASE = sys.argv[3] if len(sys.argv) > 3 else DEFAULT_OUT

JUNK = {'', 'NO', 'N/A', 'NA', 'NONE', 'NO ', '#VALUE!', '0'}


# ── 入庫單號邏輯（日期由車趟自動帶，不寫死）──
def parse_trip(trip):
    """6/26-1 -> ('20260626','-1')；6/25空調-1 -> ('20260625','aircon-1')"""
    s = str(trip or '').strip()
    m = re.match(r'(\d{1,2})/(\d{1,2})(.*)$', s)
    if m:
        mm, dd = m.group(1).zfill(2), m.group(2).zfill(2)
        suffix = re.sub(r'空調|恆溫', 'aircon', m.group(3) or '')
        return '2026' + mm + dd, suffix
    suffix = re.sub(r'\s+', '', re.sub(r'空調|恆溫', 'aircon', s))
    return '2026', suffix


def build_order_id(trip, serial):
    date, suffix = parse_trip(trip)
    oid = 'lotte' + date + suffix + serial
    oid = ''.join(c for c in oid if 0x20 <= ord(c) <= 0x7E)  # CODE128 僅 ASCII
    return oid[:30]


def clean_barcode(v):
    if v is None:
        return ''
    s = str(v).strip()
    if s.endswith('.0'):
        s = s[:-2]
    if s.upper() in JUNK:
        return ''
    return s if s.isdigit() else ''


def build_master_lookup():
    wb = openpyxl.load_workbook(MASTER, data_only=True, read_only=True)
    lookup = {}
    # 1) DSV系統 商品資料：品號0, 品名1, 條碼(單品), 箱條碼(外箱)
    dsv = wb['DSV系統 商品資料']
    drows = dsv.iter_rows(values_only=True)
    dh = next(drows)
    di = {n: i for i, n in enumerate(dh)}
    for r in drows:
        sku = r[0]
        if not sku:
            continue
        sku = str(sku).strip()
        lookup.setdefault(sku, {'single': '', 'box': '', 'is_gift': False, 'name': r[1] or ''})
        s = clean_barcode(r[di.get('條碼')]) if '條碼' in di else ''
        b = clean_barcode(r[di.get('箱條碼')]) if '箱條碼' in di else ''
        if s and not lookup[sku]['single']:
            lookup[sku]['single'] = s
        if b and not lookup[sku]['box']:
            lookup[sku]['box'] = b
    # 2) 各國/贈品分頁
    for sn in wb.sheetnames:
        if sn in ('分析', 'DSV系統 商品資料'):
            continue
        ws = wb[sn]
        rws = list(ws.iter_rows(values_only=True))
        hrow = None
        for i, r in enumerate(rws[:4]):
            if r and any(c and 'Item Code' in str(c) for c in r):
                hrow = i
                break
        if hrow is None:
            continue
        h = rws[hrow]
        ci = {}
        for i, c in enumerate(h):
            if not c:
                continue
            cs = str(c)
            if '產品條碼' in cs: ci['single'] = i
            if '外箱條碼' in cs: ci['box'] = i
            if '產品細分類' in cs: ci['cat'] = i
        for r in rws[hrow + 1:]:
            sku = r[0]
            if not sku:
                continue
            sku = str(sku).strip()
            ent = lookup.setdefault(sku, {'single': '', 'box': '', 'is_gift': False, 'name': ''})
            if 'single' in ci:
                s = clean_barcode(r[ci['single']])
                if s and not ent['single']:
                    ent['single'] = s
            if 'box' in ci:
                b = clean_barcode(r[ci['box']])
                if b and not ent['box']:
                    ent['box'] = b
            if 'cat' in ci and r[ci['cat']] and '贈品' in str(r[ci['cat']]):
                ent['is_gift'] = True
    return lookup


def is_gift(sku, name, ent):
    if ent and ent.get('is_gift'):
        return True
    if str(sku).upper().startswith('SW'):
        return True
    if name and '贈品' in str(name):
        return True
    return False


def pick_barcode(sku, name, ent):
    if not ent:
        return '', 'none'
    if is_gift(sku, name, ent):
        bc = ent['single'] or ent['box']
        return bc, 'single' if ent['single'] else ('box' if ent['box'] else 'none')
    bc = ent['box'] or ent['single']
    return bc, 'box' if ent['box'] else ('single' if ent['single'] else 'none')


def num(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def fmt_exp(v):
    return str(v or '').strip()


def main():
    if not (os.path.exists(CUST) and os.path.exists(MASTER)):
        print("❌ 找不到輸入檔")
        sys.exit(1)

    master = build_master_lookup()
    print(f"✅ 條碼主檔載入：{len(master)} 個品號\n")

    wb = openpyxl.load_workbook(CUST, data_only=True)
    ws = wb['明細']
    rows = list(ws.iter_rows(values_only=True))
    # 明細欄位 index：品號0 品名1 數量2 大5 到期日11 儲區12 Plant13 移倉14
    I_SKU, I_NAME, I_QTY, I_BIG, I_EXP, I_PLANT, I_TRIP = 0, 1, 2, 5, 11, 13, 14

    pallets = []
    no_barcode = []
    for r in rows[1:]:
        sku = r[I_SKU]
        if not sku:
            continue
        sku = str(sku).strip()
        name = r[I_NAME] or ''
        trip = str(r[I_TRIP] or '').strip()
        ent = master.get(sku)
        bc, src = pick_barcode(sku, name, ent)
        if not bc:
            no_barcode.append((sku, name, trip))
        pallets.append({
            'sourceId': '',  # 明細無儲位
            'sku': sku, 'name': name,
            'qtyBox': num(r[I_BIG]), 'qtyPcs': num(r[I_QTY]),
            'srcLoc': '', 'destLoc': '',
            'trip': trip,
            'plant': str(r[I_PLANT] or '').strip(),   # 倉別
            'expDate': fmt_exp(r[I_EXP]),
            'ean': bc, 'bc_src': src, 'is_gift': is_gift(sku, name, ent),
        })

    n = len(pallets)

    # ── 排序：車次 > 品號小到大 > 數量小到大，每車趟流水號 3 碼 ──
    pallets.sort(key=lambda p: (p['trip'], p['sku'], p['qtyPcs']))
    counters = {}
    for p in pallets:
        counters[p['trip']] = counters.get(p['trip'], 0) + 1
        p['orderId'] = build_order_id(p['trip'], str(counters[p['trip']]).zfill(3))

    # ── 報表 ──
    from collections import Counter
    trips = Counter(p['trip'] for p in pallets)
    print(f"📊 明細 {n} 個品項 → {n} 張貼紙\n")
    print("📈 各車次張數 / 入庫單號範例：")
    for t in sorted(trips):
        sample = next(p['orderId'] for p in pallets if p['trip'] == t)
        print(f"  {t}: {trips[t]} 張  (首張 {sample})")
    gifts = sum(1 for p in pallets if p['is_gift'])
    print(f"\n🎁 贈品(單品條碼)：{gifts}　📦 一般(外箱條碼)：{n-gifts}")
    if no_barcode:
        print(f"\n⚠️  無條碼 {len(no_barcode)} 張(貼紙顯示「—」)：")
        for sku, name, trip in no_barcode[:20]:
            print(f"    [{trip}] {sku}  {str(name)[:24]}")
        if len(no_barcode) > 20:
            print(f"    ...其餘 {len(no_barcode)-20} 張")

    # ── 輸出 1：v8.3 匯入格式(10 欄)──
    cols = ['sourceId','sku','name','qtyBox','qtyPcs','srcLoc','destLoc','trip','expDate','ean']
    with open(f"{OUT_BASE}_v8.3匯入.csv", 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for p in pallets:
            w.writerow({k: p[k] for k in cols})
    wb_out = openpyxl.Workbook()
    sh = wb_out.active; sh.title = '棧板資訊'
    sh.append(cols)
    for p in pallets:
        sh.append([p[k] for k in cols])
    wb_out.save(f"{OUT_BASE}_v8.3匯入.xlsx")

    # ── 輸出 2：棧板資訊(含入庫單號 + 倉別)→ 貼 Google Sheet ──
    info_cols = ['orderId','sourceId','sku','name','qtyBox','qtyPcs','destLoc','trip','plant','expDate','ean']
    with open(f"{OUT_BASE}_含入庫單號.csv", 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=info_cols)
        w.writeheader()
        for p in pallets:
            w.writerow({k: p[k] for k in info_cols})
    wb_info = openpyxl.Workbook()
    shi = wb_info.active; shi.title = '棧板資訊'
    shi.append(info_cols)
    for p in pallets:
        shi.append([p[k] for k in info_cols])
    wb_info.save(f"{OUT_BASE}_含入庫單號.xlsx")

    print(f"\n✅ 已輸出：")
    print(f"   {OUT_BASE}_v8.3匯入.xlsx/.csv  (匯入 v8.3 列印貼紙)")
    print(f"   {OUT_BASE}_含入庫單號.xlsx/.csv  (貼 Google Sheet 棧板資訊分頁)")


if __name__ == '__main__':
    main()
