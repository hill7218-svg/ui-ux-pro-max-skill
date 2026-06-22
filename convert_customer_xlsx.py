#!/usr/bin/env python3
"""
Convert customer XLSX (總表 + 工作表1) to v8.3 printer import format.

規則：
  - 每個唯一儲位 = 一板（去重）
  - 箱數 qtyBox = 該儲位「大」欄加總（客戶已換算好，單位皆為箱）
  - 件數 qtyPcs = 該儲位「數量」加總（銷售單位）
  - 條碼 ean：贈品 → 單品條碼（產品條碼）；其他 → 外箱條碼
  - 條碼來源：商品主檔 a0d64486-...xlsm
"""

import sys, os, csv, re
try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl")
    import openpyxl

# ── 重現 v8.3 列印器的入庫單號(orderId)邏輯，確保與印出貼紙一致 ──
def parse_trip(trip):
    s = str(trip or '')
    m = re.match(r'(\d{1,2})/(\d{1,2})(.*)$', s)
    if m:
        mm, dd = m.group(1).zfill(2), m.group(2).zfill(2)
        rest = re.sub(r'空調|恆溫', 'AIRCON', m.group(3) or '')
        return mm + dd, mm + dd + rest
    rest = re.sub(r'\s+', '', re.sub(r'空調|恆溫', 'AIRCON', s))
    return '', rest

def build_order_id(trip, serial):
    mmdd, code = parse_trip(trip)
    oid = 'lotte' + '2026' + mmdd + code + serial
    oid = ''.join(c for c in oid if 0x20 <= ord(c) <= 0x7E)  # CODE128 僅 ASCII
    return oid[:30]

# 用法（明天換檔案就改這裡，或用命令列參數）：
#   python3 convert_customer_xlsx.py <客戶訂單.xlsx> <商品主檔.xlsm> [輸出路徑(不含副檔名)]
# 不帶參數時用以下預設值。
DEFAULT_CUST = "/root/.claude/uploads/cc3263d7-cb46-55b2-8ea0-151cb63cd9f8/e3d77ba7-0623__1_1.xlsx"
DEFAULT_MASTER = "/root/.claude/uploads/cc3263d7-cb46-55b2-8ea0-151cb63cd9f8/a0d64486-jtu0612____________20260605__.xlsm"
DEFAULT_OUT = "/home/user/ui-ux-pro-max-skill/ALLY_LOTTE_WH/DATA/customer_converted"

CUST = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_CUST
MASTER = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_MASTER
OUT_BASE = sys.argv[3] if len(sys.argv) > 3 else DEFAULT_OUT

JUNK = {'', 'NO', 'N/A', 'NA', 'NONE', 'NO ', '#VALUE!', '0'}

def clean_barcode(v):
    """正規化條碼，去除雜訊值"""
    if v is None:
        return ''
    s = str(v).strip()
    # openpyxl 有時讀成 float，如 4903333175451.0
    if s.endswith('.0'):
        s = s[:-2]
    if s.upper() in JUNK:
        return ''
    # 只保留數字（條碼）
    return s if s.isdigit() else ''

def build_master_lookup():
    """建立 sku -> {single, box, is_gift, name}"""
    wb = openpyxl.load_workbook(MASTER, data_only=True, read_only=True)
    lookup = {}

    # 1) DSV系統 商品資料：欄位 品號0, 品名1, 條碼13(單品), 箱條碼23(外箱)
    dsv = wb['DSV系統 商品資料']
    drows = dsv.iter_rows(values_only=True)
    dh = next(drows)
    di = {n: i for i, n in enumerate(dh)}
    for r in drows:
        sku = r[0]
        if not sku:
            continue
        sku = str(sku).strip()
        lookup.setdefault(sku, {'single': '', 'box': '', 'is_gift': False, 'name': r[1] or ''})
        s = clean_barcode(r[di.get('條碼')]) if '條碼' in di else ''
        b = clean_barcode(r[di.get('箱條碼')]) if '箱條碼' in di else ''
        if s and not lookup[sku]['single']:
            lookup[sku]['single'] = s
        if b and not lookup[sku]['box']:
            lookup[sku]['box'] = b

    # 2) 各國/贈品分頁：產品條碼 / 外箱條碼 / 產品細分類(判斷贈品)
    for sn in wb.sheetnames:
        if sn in ('分析', 'DSV系統 商品資料'):
            continue
        ws = wb[sn]
        rws = list(ws.iter_rows(values_only=True))
        hrow = None
        for i, r in enumerate(rws[:4]):
            if r and any(c and 'Item Code' in str(c) for c in r):
                hrow = i
                break
        if hrow is None:
            # 贈品包材POSM 用「客戶貨號」當 key（無條碼欄，僅補贈品標記）
            continue
        h = rws[hrow]
        ci = {}
        for i, c in enumerate(h):
            if not c:
                continue
            cs = str(c)
            if '產品條碼' in cs: ci['single'] = i
            if '外箱條碼' in cs: ci['box'] = i
            if '產品細分類' in cs: ci['cat'] = i
        for r in rws[hrow + 1:]:
            sku = r[0]
            if not sku:
                continue
            sku = str(sku).strip()
            ent = lookup.setdefault(sku, {'single': '', 'box': '', 'is_gift': False, 'name': ''})
            if 'single' in ci:
                s = clean_barcode(r[ci['single']])
                if s and not ent['single']:
                    ent['single'] = s
            if 'box' in ci:
                b = clean_barcode(r[ci['box']])
                if b and not ent['box']:
                    ent['box'] = b
            if 'cat' in ci and r[ci['cat']] and '贈品' in str(r[ci['cat']]):
                ent['is_gift'] = True

    return lookup

def is_gift(sku, name, master_ent):
    if master_ent and master_ent.get('is_gift'):
        return True
    if str(sku).upper().startswith('SW'):
        return True
    if name and '贈品' in str(name):
        return True
    return False

def pick_barcode(sku, name, master_ent):
    """贈品用單品條碼；其他用外箱條碼。無對應則退而求其次。"""
    if not master_ent:
        return '', 'none'
    if is_gift(sku, name, master_ent):
        bc = master_ent['single'] or master_ent['box']
        return bc, 'single' if master_ent['single'] else ('box' if master_ent['box'] else 'none')
    else:
        bc = master_ent['box'] or master_ent['single']
        return bc, 'box' if master_ent['box'] else ('single' if master_ent['single'] else 'none')

def num(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0

def main():
    if not (os.path.exists(CUST) and os.path.exists(MASTER)):
        print("❌ 找不到輸入檔")
        sys.exit(1)

    master = build_master_lookup()
    print(f"✅ 商品主檔載入：{len(master)} 個品號\n")

    wb = openpyxl.load_workbook(CUST, data_only=True)
    ws = wb['工作表1']
    rows = list(ws.iter_rows(values_only=True))
    # 欄位 index：儲位0,品號1,品名2,數量3,單位4,...,大6,單位7,...,到期日12,移倉16
    H = {n: i for i, n in enumerate(rows[0])}
    # 「大」與其後「單位」、第一個「單位」等有重複名稱，改用固定 index
    I_LOC, I_SKU, I_NAME, I_QTY = 0, 1, 2, 3
    I_BIG, I_EXP, I_TRIP = 6, 12, 16

    pallets = {}
    no_barcode = []
    for r in rows[1:]:
        loc = r[I_LOC]
        if not loc:
            continue
        loc = str(loc).strip()
        sku = str(r[I_SKU]).strip()
        name = r[I_NAME] or ''
        ent = master.get(sku)

        if loc not in pallets:
            bc, src = pick_barcode(sku, name, ent)
            if not bc:
                no_barcode.append((loc, sku, name))
            pallets[loc] = {
                'sourceId': loc, 'sku': sku, 'name': name,
                'qtyBox': 0, 'qtyPcs': 0, 'srcLoc': '', 'destLoc': '',
                'trip': str(r[I_TRIP] or '').strip(),
                'expDate': str(r[I_EXP] or '').strip(),
                'ean': bc, 'bc_src': src, 'is_gift': is_gift(sku, name, ent),
            }
        pallets[loc]['qtyBox'] += num(r[I_BIG])   # 大欄 = 箱
        pallets[loc]['qtyPcs'] += num(r[I_QTY])   # 數量 = 銷售單位

    n = len(pallets)
    print(f"📊 工作表1 共 {len(rows)-1} 列 → 去重後 {n} 板\n")

    from collections import Counter
    trips = Counter(p['trip'] for p in pallets.values())
    print("📈 各車次板數：")
    for t, c in sorted(trips.items()):
        print(f"  {t}: {c} 張")

    gifts = sum(1 for p in pallets.values() if p['is_gift'])
    print(f"\n🎁 贈品板數（用單品條碼）：{gifts}")
    print(f"📦 一般商品板數（用外箱條碼）：{n - gifts}")
    if no_barcode:
        print(f"\n⚠️  無條碼的板（{len(no_barcode)}，貼紙顯示「—」，多為報表/信封/POSM）：")
        for loc, sku, name in no_barcode:
            print(f"    {loc}  {sku}  {str(name)[:24]}")

    # ── 產生 orderId（與 v8.3 一致：SKU字母序 + 件數小到大，每車次流水3碼）──
    ordered = sorted(pallets.values(), key=lambda x: (str(x['sku']), int(x['qtyPcs'] or 0)))
    counters = {}
    for p in ordered:
        counters[p['trip']] = counters.get(p['trip'], 0) + 1
        p['orderId'] = build_order_id(p['trip'], str(counters[p['trip']]).zfill(3))

    # 輸出（v8.3 匯入用，10 欄）
    cols = ['sourceId','sku','name','qtyBox','qtyPcs','srcLoc','destLoc','trip','expDate','ean']
    with open(f"{OUT_BASE}.csv", 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for p in ordered:
            w.writerow({k: p[k] for k in cols})

    out = openpyxl.Workbook()
    sh = out.active
    sh.title = '棧板資訊'
    sh.append(cols)
    for p in ordered:
        sh.append([p[k] for k in cols])
    out.save(f"{OUT_BASE}.xlsx")

    # 棧板資訊（貼 Google Sheet 用，orderId 為主鍵，photo.html 查商品）
    info_cols = ['orderId','sourceId','sku','name','qtyBox','qtyPcs','destLoc','trip','expDate','ean']
    with open(f"{OUT_BASE}_棧板資訊.csv", 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=info_cols)
        w.writeheader()
        for p in ordered:
            w.writerow({k: p[k] for k in info_cols})

    print(f"\n✅ 已輸出：")
    print(f"   {OUT_BASE}.csv / .xlsx（匯入 v8.3 列印貼紙）")
    print(f"   {OUT_BASE}_棧板資訊.csv（含入庫單號，貼 Google Sheet「棧板資訊」分頁）")
    print(f"📌 共 {n} 板（含正確箱數、國際條碼、入庫單號）")

if __name__ == '__main__':
    main()
